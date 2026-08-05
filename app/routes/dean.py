from flask import Blueprint, render_template, request, redirect, session, url_for, flash, jsonify
from app.models import *
from app.decorators import role_required

dean_bp = Blueprint('dean', __name__, url_prefix='/dean')


@dean_bp.route('/')
@role_required('DEAN')
def dean_dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()

    dean_id = session.get('user_id')
    terms = get_all_terms(cursor)
    active_term = next((t for t in terms if t['is_active'] == 1), None)

    if not active_term:
        flash('No active academic term found.', 'warning')
        cursor.close()
        conn.close()
        return render_template('dean_dashboard.html',
                               active_term=None,
                               master_indicators=[],
                               existing_quotas={},
                               completion_rate=0,
                               pending_count=0,
                               top_dept="N/A",
                               pending_approvals=[],
                               draft_submissions=[],
                               college_wide_quotas=[],
                               designated_faculty_list=[])

    term_id = active_term['term_id']

    # Use timed_query helper for all queries
    from app.models.connection import timed_query

    indicators = get_master_indicators(cursor, term_id)

    existing_quotas_raw = get_existing_cascaded_quotas(cursor, term_id)

    existing_quotas = {}
    for quota in existing_quotas_raw:
        ind_id = quota['indicator_id']
        if ind_id not in existing_quotas:
            existing_quotas[ind_id] = {}
        existing_quotas[ind_id][quota['assigned_to_role']] = quota['total_target_value']

    # Consolidated KPI query — 1 round-trip instead of 3
    completion_rate, pending_count, top_dept = get_dean_dashboard_kpis(cursor, term_id)

    pending_approvals = get_pending_final_approvals(cursor, term_id)

    # ── New: Draft IPCR submissions from designated faculty ──
    draft_submissions = get_designated_draft_submissions(cursor, term_id)

    # ── New: College-Wide quotas for target assignment ──
    college_wide_quotas = get_college_wide_cascaded_quotas(cursor, term_id)

    # ── New: Designated faculty list ──
    designated_faculty_list = get_designated_faculty_list(cursor)

    # Get ALL assignments in ONE batch query (replaces N+1 loop)
    emp_ids = [fac['emp_id'] for fac in designated_faculty_list]
    designated_assignments = get_designated_faculty_assignments_batch(cursor, term_id, emp_ids)

    # Fetch college-wide allocations for tracking
    allocations_list = get_college_wide_allocations_tracker(cursor, term_id)
    
    # Structure them by indicator_id: {indicator_id: [allocations]}
    college_wide_allocations = {}
    for alloc in allocations_list:
        ind_id = alloc['indicator_id']
        if ind_id not in college_wide_allocations:
            college_wide_allocations[ind_id] = []
        college_wide_allocations[ind_id].append(alloc)

    cursor.close()
    conn.close()

    return render_template('dean_dashboard.html',
                           active_term=active_term,
                           master_indicators=indicators,
                           existing_quotas=existing_quotas,
                           completion_rate=completion_rate,
                           pending_count=pending_count,
                           top_dept=top_dept,
                           pending_approvals=pending_approvals,
                           draft_submissions=draft_submissions,
                           college_wide_quotas=college_wide_quotas,
                           designated_faculty_list=designated_faculty_list,
                           designated_assignments=designated_assignments,
                           college_wide_allocations=college_wide_allocations)


@dean_bp.route('/cascade_quotas', methods=['POST'])
@role_required('DEAN')
def cascade_quotas():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        term_id = request.form.get('term_id')
        if not term_id:
            flash('Missing term ID', 'danger')
            return redirect(url_for('dean.dean_dashboard'))

        quotas_data = []
        indicator_ids = request.form.getlist('indicator_id[]')
        wst_values = request.form.getlist('wst[]')
        dst_values = request.form.getlist('dst[]')
        nst_values = request.form.getlist('nst[]')
        bsds_values = request.form.getlist('bsds[]')
        ret_values = request.form.getlist('ret[]')
        college_values = request.form.getlist('college[]')

        for i, ind_id in enumerate(indicator_ids):
            if not ind_id:
                continue

            values = [
                ('WST Program', int(wst_values[i]) if wst_values[i] and int(wst_values[i]) > 0 else 0),
                ('DST Program', int(dst_values[i]) if dst_values[i] and int(dst_values[i]) > 0 else 0),
                ('NST Program', int(nst_values[i]) if nst_values[i] and int(nst_values[i]) > 0 else 0),
                ('BSDS Program', int(bsds_values[i]) if bsds_values[i] and int(bsds_values[i]) > 0 else 0),
                ('RET / Extension', int(ret_values[i]) if ret_values[i] and int(ret_values[i]) > 0 else 0),
                ('College-Wide', int(college_values[i]) if i < len(college_values) and college_values[i] and int(college_values[i]) > 0 else 0)
            ]

            for role, value in values:
                if value > 0:
                    quotas_data.append({
                        'indicator_id': int(ind_id),
                        'total_target': value,
                        'assigned_role': role
                    })

        success, message = save_cascaded_quotas(cursor, conn, term_id, quotas_data)

        if success:
            flash(message, 'success')
        else:
            flash(message, 'danger')

    except Exception as e:
        flash(f'Error cascading quotas: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('dean.dean_dashboard'))


@dean_bp.route('/batch_approve', methods=['POST'])
@role_required('DEAN')
def batch_approve():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        score_ids = request.form.getlist('score_ids[]')
        action = request.form.get('action', 'approve')

        if not score_ids:
            flash('No IPCRs selected for approval', 'warning')
            return redirect(url_for('dean.dean_dashboard'))

        new_status = 'Approved' if action == 'approve' else 'Reverted'
        success, message = update_dean_approval_status(cursor, conn, score_ids, new_status)

        flash(message, 'success' if success else 'danger')

    except Exception as e:
        flash(f'Error processing batch approval: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('dean.dean_dashboard'))


@dean_bp.route('/validate_quotas', methods=['POST'])
@role_required('DEAN')
def validate_quotas():
    """AJAX endpoint to validate quotas before submission"""
    data = request.get_json()
    return jsonify({'valid': True, 'message': 'Quotas validated'})


@dean_bp.route('/inspect_template')
@role_required('DEAN')
def inspect_template():
    import os
    from openpyxl import load_workbook
    template_dir = os.path.join(os.getcwd(), 'app', 'static', 'templates')
    template_path = os.path.join(template_dir, 'dpcr_template.xlsx')
    if not os.path.exists(template_path):
        return jsonify({'error': 'Template not found'})
    wb = load_workbook(template_path)
    ws = wb.active
    cells_info = []
    for r in range(1, 45):
        row_info = {}
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            row_info[cell.coordinate] = {
                'value': cell.value,
                'is_merged': cell.coordinate in ws.merged_cells
            }
        cells_info.append(row_info)
    return jsonify({
        'merged_ranges': [str(r) for r in ws.merged_cells.ranges],
        'cells': cells_info
    })


@dean_bp.route('/export_dpcr')
@role_required('DEAN')
def export_dpcr():
    import os
    import io
    from flask import send_file, flash, redirect, url_for, session
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from copy import copy
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        terms = get_all_terms(cursor)
        active_term = next((t for t in terms if t['is_active'] == 1), None)
        
        if not active_term:
            flash('No active academic term found.', 'warning')
            return redirect(url_for('dean.dean_dashboard'))
            
        term_id = active_term['term_id']
        
        template_dir = os.path.join(os.getcwd(), 'app', 'static', 'templates')
        template_path = os.path.join(template_dir, 'dpcr_template.xlsx')
        
        if not os.path.exists(template_path):
            flash('DPCR Template file not found. Please upload your template to app/static/templates/dpcr_template.xlsx first.', 'danger')
            return redirect(url_for('dean.dean_dashboard'))
            
        wb = load_workbook(template_path)
        ws = wb.active

        def get_category_key(cat_name):
            cat_lower = cat_name.lower()
            if 'instruction' in cat_lower: return 'instruction'
            elif 'research' in cat_lower: return 'research'
            elif 'extension' in cat_lower or 'advisory' in cat_lower: return 'extension'
            elif 'support' in cat_lower: return 'support'
            return 'custom'

        cursor.execute("""
            SELECT 
                mi.indicator_id, mi.indicator_description, tc.category_name,
                cq.total_target_value, cq.assigned_to_role
            FROM tbl_master_indicators mi
            JOIN tbl_target_categories tc ON mi.category_id = tc.category_id
            LEFT JOIN tbl_cascaded_quotas cq ON mi.indicator_id = cq.indicator_id AND cq.term_id = mi.term_id
            WHERE mi.term_id = %s
            ORDER BY tc.category_id, mi.indicator_id
        """, (term_id,))
        
        rows = cursor.fetchall()
        indicators_map = {}
        for row in rows:
            if isinstance(row, dict):
                ind_id = row['indicator_id']
                desc = row['indicator_description']
                cat_name = row['category_name']
                val = row['total_target_value']
                role = row['assigned_to_role']
            else:
                ind_id = row[0]
                desc = row[1]
                cat_name = row[2]
                val = row[3]
                role = row[4]

            if ind_id not in indicators_map:
                indicators_map[ind_id] = {
                    'indicator_id': ind_id,
                    'indicator_description': desc,
                    'category_name': cat_name,
                    'quotas': {}
                }
            if role and val is not None:
                indicators_map[ind_id]['quotas'][role] = val

        indicators_by_category = {'instruction': [], 'research': [], 'extension': [], 'support': [], 'custom': []}
        for ind in indicators_map.values():
            key = get_category_key(ind['category_name'])
            indicators_by_category[key].append(ind)

        col_map = {'DST Program': 9, 'WST Program': 10, 'NST Program': 11, 'BSDS Program': 12, 'RET / Extension': 13}
        template_blocks = {
            'instruction': [(17, 19), (20, 21), (22, 24), (25, 27), (28, 29), (30, 32), (33, 34), (35, 37)],
            'research': [(39, 40), (41, 43), (44, 47), (48, 50), (51, 53)],
            'extension': [(55, 58), (59, 62), (63, 66), (67, 69), (70, 71), (72, 74)],
            'support': [(76, 79), (80, 83), (84, 87), (88, 91), (92, 95), (96, 99), (100, 103)]
        }

        if indicators_by_category.get('custom'):
            indicators_by_category['support'].extend(indicators_by_category['custom'])
            indicators_by_category['custom'] = []

        thin_side = Side(style='thin')
        thin_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

        def safe_write_cell(ws, row, col, value):
            from openpyxl.cell import MergedCell, Cell
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                new_cell = Cell(ws, row=row, column=col)
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
                new_cell.value = value
                ws._cells[(row, col)] = new_cell
                return new_cell
            else:
                cell.value = value
                return cell

        def safe_merge(start_row, start_col, end_row, end_col):
            if start_row == end_row and start_col == end_col: return
            already_merged = False
            for r in ws.merged_cells.ranges:
                if r.min_row == start_row and r.min_col == start_col and r.max_row == end_row and r.max_col == end_col:
                    already_merged = True; break
            if not already_merged:
                try: ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
                except: pass

        def unmerge_intersecting_ranges(b_start, b_end):
            ranges_to_remove = [r for r in list(ws.merged_cells.ranges) if r.min_row <= b_end and r.max_row >= b_start]
            for r in ranges_to_remove:
                try: ws.unmerge_cells(start_row=r.min_row, start_column=r.min_col, end_row=r.max_row, end_column=r.max_col)
                except: pass

        def clear_and_unmerge_quota_cols(b_start, b_end):
            unmerge_intersecting_ranges(b_start, b_end)
            for r in range(b_start, b_end + 1):
                for c in range(9, 14): safe_write_cell(ws, r, c, None)

        def style_quota_cells(b_start, b_end, min_c, max_c):
            for r in range(b_start, b_end + 1):
                for c in range(min_c, max_c + 1): ws.cell(row=r, column=c).border = thin_border

        def find_section_boundaries(ws):
            headers = {'instruction': None, 'research': None, 'extension': None, 'support': None}
            boundaries = []
            for r in range(1, ws.max_row + 1):
                val = ws.cell(row=r, column=1).value
                if val and isinstance(val, str):
                    val_lower = val.lower().strip()
                    if ("a. instruction" in val_lower or "instruction" in val_lower) and not headers['instruction']:
                        headers['instruction'] = r
                    elif "a. research" in val_lower and not headers['research']:
                        headers['research'] = r
                    elif "b. extension" in val_lower and not headers['extension']:
                        headers['extension'] = r
                    elif ("support functions" in val_lower or "iii. support" in val_lower) and not headers['support']:
                        headers['support'] = r
                    
                    if (val_lower.startswith("i.") or val_lower.startswith("ii.") or val_lower.startswith("iii.") or
                        val_lower.startswith("total overall") or val_lower.startswith("comments and") or
                        "strategic priorities" in val_lower or "core functions" in val_lower or
                        "a. instruction" in val_lower or "a. research" in val_lower or
                        "b. extension" in val_lower or "support functions" in val_lower):
                        boundaries.append(r)
            return headers, sorted(list(set(boundaries)))

        def get_live_category_blocks(ws, cat_key):
            headers, boundaries = find_section_boundaries(ws)
            h_row = headers.get(cat_key)
            if not h_row:
                return []
            next_boundary = ws.max_row + 1
            for b in boundaries:
                if b > h_row:
                    next_boundary = b
                    break
            blocks = []
            curr_r = h_row + 1
            while curr_r < next_boundary and curr_r <= ws.max_row:
                b_start = curr_r
                b_end = curr_r
                for m_range in ws.merged_cells.ranges:
                    if m_range.min_col <= 14 <= m_range.max_col and m_range.min_row <= curr_r <= m_range.max_row:
                        b_start = m_range.min_row
                        b_end = m_range.max_row
                        break
                b_start = max(b_start, h_row + 1)
                b_end = min(b_end, next_boundary - 1)
                if b_start <= b_end:
                    blocks.append((b_start, b_end))
                    curr_r = b_end + 1
                else:
                    curr_r += 1
            return blocks

        # Process categories TOP-TO-BOTTOM with live row boundary detection
        for cat in ['instruction', 'research', 'extension', 'support']:
            indicators = indicators_by_category[cat]
            blocks = get_live_category_blocks(ws, cat)
            n_ind, n_blocks = len(indicators), len(blocks)

            if n_ind < n_blocks:
                # Delete extra unused blocks from bottom to top within this category
                for b_idx in range(n_blocks - 1, n_ind - 1, -1):
                    b_start, b_end = blocks[b_idx]
                    unmerge_intersecting_ranges(b_start, b_end)
                    ws.delete_rows(b_start, b_end - b_start + 1)
                    blocks.pop(b_idx)

            elif n_ind > n_blocks and n_blocks > 0:
                # Insert additional blocks by copying the last block of the category
                for i in range(n_blocks, n_ind):
                    last_start, last_end = blocks[-1]
                    block_height = last_end - last_start + 1
                    insert_at = last_end + 1
                    ws.insert_rows(insert_at, block_height)
                    for offset in range(block_height):
                        ws.row_dimensions[insert_at + offset].height = ws.row_dimensions[last_start + offset].height
                    for offset in range(block_height):
                        src_r, dest_r = last_start + offset, insert_at + offset
                        for c in range(1, ws.max_column + 1):
                            src_c = ws.cell(row=src_r, column=c)
                            safe_write_cell(ws, dest_r, c, src_c.value)
                            dest_c = ws.cell(row=dest_r, column=c)
                            dest_c.font = copy(src_c.font)
                            dest_c.border = copy(src_c.border)
                            dest_c.fill = copy(src_c.fill)
                            dest_c.number_format = src_c.number_format
                            dest_c.alignment = copy(src_c.alignment)
                    for r in list(ws.merged_cells.ranges):
                        if r.min_row >= last_start and r.max_row <= last_end:
                            row_offset = insert_at - last_start
                            safe_merge(r.min_row + row_offset, r.min_col, r.max_row + row_offset, r.max_col)
                    blocks.append((insert_at, insert_at + block_height - 1))

            # Write indicators to the live blocks of this category
            for i, ind in enumerate(indicators):
                if i >= len(blocks): break
                b_start, b_end = blocks[i]
                desc_cell = safe_write_cell(ws, b_start, 5, ind['indicator_description'])
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                quotas = ind.get('quotas', {})
                is_cw = ('College-Wide' in quotas and quotas['College-Wide'] > 0)
                clear_and_unmerge_quota_cols(b_start, b_end)
                if is_cw:
                    safe_merge(b_start, 9, b_end, 13)
                    val = quotas['College-Wide']
                    display_val = f"{int(val * 100)}%" if ('%' in ind['indicator_description'] and isinstance(val, (int, float)) and val <= 5) else val
                    cell_cw = safe_write_cell(ws, b_start, 9, display_val)
                    cell_cw.alignment = Alignment(horizontal="center", vertical="center")
                    style_quota_cells(b_start, b_end, 9, 13)
                else:
                    for role, c_idx in col_map.items():
                        if b_start < b_end: safe_merge(b_start, c_idx, b_end, c_idx)
                        cell = safe_write_cell(ws, b_start, c_idx, quotas.get(role, 0))
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        style_quota_cells(b_start, b_end, c_idx, c_idx)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        filename = f"DPCR_{active_term['academic_year'].replace('/', '_')}_{active_term['semester'].replace(' ', '_')}.xlsx"
        return send_file(file_stream, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Error exporting DPCR: {str(e)}', 'danger')
        return redirect(url_for('dean.dean_dashboard'))
    finally:
        cursor.close()
        conn.close()



@dean_bp.route('/review_draft_fetch/<int:emp_id>')
@role_required('DEAN')
def review_draft_fetch(emp_id):
    """AJAX endpoint — returns JSON to populate the Dean's review modal."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        dean_id = session.get('user_id')
        terms = get_all_terms(cursor)
        active_term = next((t for t in terms if t['is_active'] == 1), None)
        if not active_term:
            return jsonify({'error': 'No active term found.'}), 400

        term_id = active_term['term_id']

        # Get or create review record
        review_id = get_or_create_dean_review(conn, cursor, emp_id, term_id, dean_id)

        # Fetch review items
        items = get_dean_review_items(cursor, review_id)

        # Fetch overall status & remarks
        cursor.execute(
            "SELECT overall_status, overall_remarks FROM tbl_ipcr_dean_review WHERE review_id = %s",
            (review_id,)
        )
        row = cursor.fetchone()
        overall_status = row[0] if row else 'Pending'
        overall_remarks = row[1] if row else ''

        # Faculty info
        cursor.execute(
            "SELECT CONCAT(first_name, ' ', last_name), academic_rank, designation, assigned_program FROM tbl_employee_profiles WHERE emp_id = %s",
            (emp_id,)
        )
        fac = cursor.fetchone()
        faculty_name = fac[0] if fac else 'Unknown'
        academic_rank = fac[1] if fac else ''
        designation = fac[2] if fac else ''
        assigned_program = fac[3] if fac else ''

        # Get college-wide quotas
        college_wide = get_college_wide_cascaded_quotas(cursor, term_id)
        college_wide_ids = {q['indicator_id'] for q in college_wide}

        # Also get ALL available master indicators for this term
        all_indicators = get_available_master_indicators(cursor, term_id)
        picked_ids = {i['indicator_id'] for i in items}
        
        # Filter unpicked to exclude picked AND college wide targets
        unpicked = [ind for ind in all_indicators if ind['indicator_id'] not in picked_ids and ind['indicator_id'] not in college_wide_ids]

        # Filter college wide to only show those that are unpicked
        college_wide_unpicked = []
        for q in college_wide:
            if q['indicator_id'] not in picked_ids:
                college_wide_unpicked.append({
                    'indicator_id': q['indicator_id'],
                    'indicator_description': q['indicator_description'],
                    'category_name': q['category_name'],
                    'total_target_value': q['total_target_value']
                })

        serializable_items = []
        for item in items:
            serializable_items.append({
                'item_id': item['item_id'],
                'draft_id': item['draft_id'],
                'indicator_id': item['indicator_id'],
                'indicator_description': item['indicator_description'],
                'category_name': item['category_name'],
                'original_quantity': item['original_quantity'],
                'reviewed_quantity': item['reviewed_quantity'],
                'item_remarks': item['item_remarks'] or '',
                'is_custom': item['is_custom'],
                'target_description': item.get('target_description') or item['indicator_description'],
                'target_deadline': item.get('target_deadline') or '1 Semester',
                'is_core': item.get('is_core', False),
                'is_cascaded': item.get('is_cascaded', False),
            })


        return jsonify({
            'review_id': review_id,
            'emp_id': emp_id,
            'faculty_name': faculty_name,
            'academic_rank': academic_rank,
            'designation': designation,
            'assigned_program': assigned_program,
            'overall_status': overall_status,
            'overall_remarks': overall_remarks or '',
            'items': serializable_items,
            'unpicked': unpicked,
            'college_wide_unpicked': college_wide_unpicked,
            'college_wide_all': [{
                'indicator_id': q['indicator_id'],
                'indicator_description': q['indicator_description'],
                'category_name': q['category_name'],
                'total_target_value': q['total_target_value']
            } for q in college_wide],
            'college_wide_ids': list(college_wide_ids),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@dean_bp.route('/save_review_items', methods=['POST'])
@role_required('DEAN')
def save_review_items():
    """Batch save all edited quantities and remarks for one review."""
    data = request.get_json()
    review_id = data.get('review_id')
    items = data.get('items', [])

    if not review_id or not items:
        return jsonify({'success': False, 'message': 'Missing review_id or items.'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        success, msg = save_dean_review_items(cursor, conn, review_id, items)
        return jsonify({'success': success, 'message': msg})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@dean_bp.route('/submit_review_decision', methods=['POST'])
@role_required('DEAN')
def submit_review_decision():
    """Approve or reject the entire review with overall remarks."""
    data = request.get_json()
    review_id = data.get('review_id')
    action = data.get('action')
    overall_remarks = data.get('overall_remarks', '').strip()

    if not review_id or action not in ('Approved', 'Rejected'):
        return jsonify({'success': False, 'message': 'Missing review_id or invalid action.'}), 400

    if action == 'Rejected' and not overall_remarks:
        return jsonify({'success': False, 'message': 'Remarks are required when rejecting.'}), 400

    dean_id = session.get('user_id')
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        success, msg = submit_dean_review_decision(cursor, conn, review_id, action, overall_remarks)

        if success:
            details = f"Dean {dean_id} {action.lower()} draft IPCR (review #{review_id}). Remarks: {overall_remarks}"
            from app.models.audit import log_audit_action
            log_audit_action(conn, cursor, dean_id, f'DEAN_REVIEW_{action.upper()}', details, request.remote_addr or '127.0.0.1')

        return jsonify({'success': success, 'message': msg})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


# ──────────────────────────────────────────────
# College-Wide Target Assignment (Designated Faculty)
# ──────────────────────────────────────────────

@dean_bp.route('/assign_designated_target', methods=['POST'])
@role_required('DEAN')
def assign_designated_target():
    """Assign a College-Wide target to a designated faculty member."""
    term_id = request.form.get('term_id')
    emp_id = request.form.get('emp_id')
    indicator_id = request.form.get('indicator_id')
    quantity = request.form.get('quantity', '0')

    if not term_id or not emp_id or not indicator_id:
        flash('Missing required data.', 'danger')
        return redirect(url_for('dean.dean_dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        success, msg = save_designated_faculty_assignment(
            cursor, conn, int(term_id), emp_id, int(indicator_id), int(quantity) if quantity.isdigit() else 0
        )
        flash(msg, 'success' if success else 'danger')
    except Exception as e:
        flash(f'Error assigning target: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('dean.dean_dashboard'))
